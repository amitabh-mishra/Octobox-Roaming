import json, os, csv, xlrd, openpyxl
import pandas as pd
from openpyxl import Workbook, load_workbook
from os import sys

##################################################################################################################################

# cwd = os.getcwd()
# files = os.listdir(cwd)
# print("Files in %r:configFile.close() %s" % (cwd, files))
# /Users/Amitabh.Mishra/Documents/Octobox/Mesh Script.py 

# Linux paths: 
# "playback_request.json"
# "csv/sweep_5000_0_3_90.csv"

# os.system("cd ~/Desktop/quadatten_playback")

##################################################################################################################################

# SET FILE PATHS
playback_request_filepath = "/Users/Amitabh.Mishra/Documents/Octobox/playback_request.json"
sweep_xlsx_filepath = "/Users/Amitabh.Mishra/Documents/Octobox/csv/sweep.xlsx"

# LOAD SWEEP EXCEL FILE AND SELECT
sweep_workbook = load_workbook(sweep_xlsx_filepath)
sweep_sheet = sweep_workbook.active

create_sweep = input("Would you like to create a config (Y/N)? ")

while(create_sweep == "Y" or create_sweep == "y"):
    
    # ASK USER FOR DWELL, START, STEP, AND STOP VALUES
    user_dwell_value = input("Enter dwell value: ")
    user_start_value = input("Enter start value: ")
    user_step_value = input("Enter step value: ")
    user_stop_value = input("Enter stop value: ")

    print("")

    # UPDATE SWEEP SHEET WITH USER VALUES
    sweep_sheet["B1"] = float(user_dwell_value)
    sweep_sheet["B2"] = float(user_start_value)
    sweep_sheet["B3"] = float(user_step_value)
    sweep_sheet["B4"] = float(user_stop_value)

    # SAVE WORKBOOK
    sweep_workbook.save(sweep_xlsx_filepath)

    # CREATE A NEW EXCEL SHEET
    sweep_config_fileName = input("Please the name the file: ")
    user_sweep_config = openpyxl.Workbook()
    user_sweep_config.save("/Users/Amitabh.Mishra/Documents/Octobox/csv/" + sweep_config_fileName + ".xlsx")
    user_sweep_xlsx_filepath = "/Users/Amitabh.Mishra/Documents/Octobox/csv/" + sweep_config_fileName + ".xlsx"

    # COPY CONFIG FROM SWEEP FILE TO USER SWEEP FILE
    sweep_workbook = load_workbook(sweep_xlsx_filepath, data_only=True)
    user_sweep_workbook = load_workbook(user_sweep_xlsx_filepath)
    source = sweep_workbook.get_sheet_by_name("csv")
    destination = user_sweep_workbook.get_sheet_by_name("csv")

    read_file = pd.read_excel(sweep_xlsx_filepath, sheet_name = "csv", index_col=None)
    read_file.to_csv(user_sweep_xlsx_filepath, encoding = "utf-8")

    for i in range(1, source.max_row + 1):
        for j in range(1, source.max_column + 1):
            destination.cell(row = i, column = j).value = source.cell(row = i, column = j).value
            
    # SAVE WORKBOOKS
    sweep_workbook.save(sweep_xlsx_filepath)
    user_sweep_workbook.save(user_sweep_xlsx_filepath)

    create_sweep = input("Would you like to create another config (Y/N)? ")

##################################################################################################################################

edit_playback_file = input("Would you like to edit a playback request file (Y/N)?")

while(edit_playback_file == "Y" or edit_playback_file == "y"):

    open_playback_file = input("Enter playback_request file name: ")

    open_request_filepath = "/Users/Amitabh.Mishra/Documents/Octobox/" + open_playback_file + ".json"

    device_ip_address = input("Enter device ip address: ")
    # edit_loopback_delay = input("Enter loopback delay: ")
    # edit_loopback_mode = input("Enter loopback mode: ")
    edit_sweep_file = input("Enter sweep (config) file name: ")
    # edit_loopback = input("Enter loopback value: ")

    # OPEN PLAYBACK_REQUEST.JSON FILE AND CONFIGURE
    with open(open_request_filepath, "r+") as configFile:
        configData = json.load(configFile)

        configData["device_ip_addr"] = "'" + device_ip_address + "'"
        configData["playback_filename"] = ["'" + 'csv/' + edit_sweep_file + "'"]

        configFile.seek(0)
        json.dump(configData, configFile, indent=4)
        configFile.truncate()

    configFile.close()

    edit_playback_file = input("Would you like to edit another playback request file (Y/N)?")

os.system("node main playback_request.json")

# To run multiple, add "&" between each system call
# os.system(node main playback_request.json & node main playback playback_request_2.json)

##################################################################################################################################

#     "device_ip_addr":       "10.19.4.6",
#     "loopback_delay_usec":  0,
#     "loopback_mode":        "disabled",
#     "playback_filename":    ["csv/Transition_FastMobility.csv"],
#     "loopback":				15

##################################################################################################################################


